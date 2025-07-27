import json
from flask import Flask, render_template, request, send_from_directory, jsonify
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import cm
import os

app = Flask(__name__)

# On serverless platforms like Vercel or Cloudflare Pages, use /tmp for writable files
if os.environ.get('VERCEL') or os.environ.get('CF_PAGES'):
    writable_dir = '/tmp'
else:
    # For local development, use the project directory
    writable_dir = os.path.abspath(os.path.dirname(__file__))

# Define paths for instance (database) and output (generated files)
instance_path = os.path.join(writable_dir, 'instance')
output_path = os.path.join(writable_dir, 'output')

# Create these directories if they don't exist
os.makedirs(instance_path, exist_ok=True)
os.makedirs(output_path, exist_ok=True)

# The base directory for reading non-writable files like fonts
basedir = os.path.abspath(os.path.dirname(__file__))

# Use DATABASE_URL from environment variables if set, otherwise use the writable path
default_db_path = os.path.join(instance_path, 'quotations.db')
database_url = os.environ.get('DATABASE_URL', f'sqlite:///{default_db_path}')

# Vercel's Postgres uses postgres:// which SQLAlchemy needs as postgresql://
if database_url and database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)

app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

class Quotation(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    company_name = db.Column(db.String(100))
    company_address = db.Column(db.String(200))
    company_phone = db.Column(db.String(20))
    company_email = db.Column(db.String(100))
    client_name = db.Column(db.String(100))
    client_address = db.Column(db.String(200))
    quotation_no = db.Column(db.String(50))
    date = db.Column(db.String(50))
    items = db.Column(db.Text)  # Store items as a JSON string
    received = db.Column(db.Float)
    deposit_info = db.Column(db.Text)

    def to_dict(self):
        return {
            'id': self.id,
            'company_name': self.company_name,
            'company_address': self.company_address,
            'company_phone': self.company_phone,
            'company_email': self.company_email,
            'client_name': self.client_name,
            'client_address': self.client_address,
            'quotation_no': self.quotation_no,
            'date': self.date,
            'items': self.items, # Keep as JSON string for loading
            'received': self.received,
            'deposit_info': self.deposit_info
        }

# Register Chinese font
# The font is expected to be in the 'static' directory relative to the project root
pdfmetrics.registerFont(TTFont('SimSun', os.path.join(basedir, 'static', 'SimSun.ttf')))

with app.app_context():
    db.create_all()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/quotations')
def get_quotations():
    quotations = Quotation.query.order_by(Quotation.date.desc()).all()
    return jsonify([q.to_dict() for q in quotations])

@app.route('/api/save_quotation', methods=['POST'])
def save_quotation():
    data = request.get_json()
    quotation_id = data.get('id')

    items_json = json.dumps(data.get('items', []))

    if quotation_id:
        # Update existing quotation
        quotation = Quotation.query.get(quotation_id)
        if not quotation:
            return jsonify({'success': False, 'message': 'Quotation not found'}), 404
    else:
        # Create new quotation
        quotation = Quotation()

    quotation.company_name = data.get('company_name')
    quotation.company_address = data.get('company_address')
    quotation.company_phone = data.get('company_phone')
    quotation.company_email = data.get('company_email')
    quotation.client_name = data.get('client_name')
    quotation.client_address = data.get('client_address')
    quotation.quotation_no = data.get('quotation_no')
    quotation.date = data.get('date')
    quotation.received = data.get('received')
    quotation.deposit_info = data.get('deposit_info')
    quotation.items = items_json

    if not quotation_id:
        db.session.add(quotation)
        
    db.session.commit()
    return jsonify({'success': True, 'id': quotation.id, 'message': 'Quotation saved successfully'})

@app.route('/api/get_quotation/<int:id>')
def get_quotation(id):
    quotation = Quotation.query.get(id)
    if quotation:
        return jsonify(quotation.to_dict())
    return jsonify({'error': 'Quotation not found'}), 404

@app.route('/api/delete_quotation/<int:id>', methods=['DELETE'])
def delete_quotation(id):
    quotation = Quotation.query.get(id)
    if quotation:
        db.session.delete(quotation)
        db.session.commit()
        return jsonify({'success': True})
    return jsonify({'error': 'Quotation not found'}), 404


@app.route('/generate', methods=['POST'])
def generate():
    # Collect data from form
    data = {
        'company_name': request.form['company_name'],
        'company_address': request.form['company_address'],
        'company_phone': request.form['company_phone'],
        'company_email': request.form['company_email'],
        'quotation_no': request.form['quotation_no'],
        'date': request.form['date'],
        'client_name': request.form['client_name'],
        'client_address': request.form['client_address'],
        'received': float(request.form.get('received', '0') or '0'),
        'deposit_info': request.form.get('deposit_info', ''),
        'items': []
    }

    item_names = request.form.getlist('item_name[]')
    quantities = request.form.getlist('quantity[]')
    prices = request.form.getlist('price[]')

    total_amount = 0
    for i in range(len(item_names)):
        if not item_names[i]: continue
        quantity = float(quantities[i])
        price = float(prices[i])
        amount = quantity * price
        data['items'].append({'name': item_names[i], 'quantity': quantity, 'price': price, 'amount': amount})
        total_amount += amount
    
    data['total_amount'] = total_amount
    data['balance'] = total_amount - data['received']
    items = data.pop('items')

    return render_template('preview.html', data=data, items=items)

@app.route('/create_files', methods=['POST'])
def create_files():
    # Get data from hidden fields
    company_name = request.form['company_name']
    company_address = request.form['company_address']
    company_phone = request.form['company_phone']
    company_email = request.form['company_email']
    quotation_no = request.form['quotation_no']
    date = request.form['date']
    client_name = request.form['client_name']
    client_address = request.form['client_address']
    received = float(request.form.get('received', '0') or '0')
    deposit_info = request.form.get('deposit_info', '')

    item_names = request.form.getlist('item_name[]')
    quantities = request.form.getlist('quantity[]')
    prices = request.form.getlist('price[]')
    items = []
    total_amount = 0
    for i in range(len(item_names)):
        if not item_names[i]: continue
        quantity = float(quantities[i])
        price = float(prices[i])
        amount = quantity * price
        items.append({'name': item_names[i], 'quantity': quantity, 'price': price, 'amount': amount})
        total_amount += amount

    balance = total_amount - received

    # Generate Excel and PDF
    excel_filename = generate_excel(company_name, company_address, company_phone, company_email, quotation_no, date, client_name, client_address, items, total_amount, received, balance, deposit_info)
    pdf_filename = generate_pdf(company_name, company_address, company_phone, company_email, quotation_no, date, client_name, client_address, items, total_amount, received, balance, deposit_info)

    return render_template('result.html', excel_file=excel_filename, pdf_file=pdf_filename)

@app.route('/download/<filename>')
def download_file(filename):
    return send_from_directory(output_path, filename, as_attachment=True)

def generate_excel(company_name, company_address, company_phone, company_email, quotation_no, date, client_name, client_address, items, total_amount, received, balance, deposit_info):
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"

    # Set styles
    font_bold = Font(bold=True)
    align_center = Alignment(horizontal='center', vertical='center')
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Company Info
    ws.merge_cells('A1:E1')
    ws['A1'] = company_name
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = align_center

    ws.merge_cells('A2:E2')
    ws['A2'] = company_address
    ws['A2'].alignment = align_center

    ws.merge_cells('A3:E3')
    ws['A3'] = f"Tel: {company_phone} Email: {company_email}"
    ws['A3'].alignment = align_center

    # Quotation Title
    ws.merge_cells('A4:E4')
    ws['A4'] = "報價單"
    ws['A4'].font = Font(size=14, bold=True)
    ws['A4'].alignment = align_center

    # Client Info
    ws['A6'] = "客戶名稱:"
    ws['B6'] = client_name
    ws['D6'] = "報價單號碼:"
    ws['E6'] = quotation_no

    ws['A7'] = "客戶地址:"
    ws['B7'] = client_address
    ws['D7'] = "日期:"
    ws['E7'] = date

    # Table Header
    headers = ["項目", "數量", "單價", "金額"]
    ws.append(headers)
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}9'].font = font_bold
        ws[f'{col}9'].alignment = align_center
        ws[f'{col}9'].border = border_thin

    # Items
    row = 10
    for item in items:
        ws.cell(row, 1, item['name'])
        ws.cell(row, 2, item['quantity'])
        ws.cell(row, 3, item['price'])
        ws.cell(row, 4, item['amount'])
        for col in range(1, 5):
            ws.cell(row, col).border = border_thin
        row += 1

    # Total
    ws.cell(row, 3, "總計:").font = font_bold
    ws.cell(row, 4, total_amount).font = font_bold
    ws.cell(row, 3).border = border_thin
    ws.cell(row, 4).border = border_thin
    row += 1
    ws.cell(row, 3, "已收訂金:").font = font_bold
    ws.cell(row, 4, received).font = font_bold
    ws.cell(row, 3).border = border_thin
    ws.cell(row, 4).border = border_thin
    row += 1
    ws.cell(row, 3, "餘額:").font = font_bold
    ws.cell(row, 4, balance).font = font_bold
    ws.cell(row, 3).border = border_thin
    ws.cell(row, 4).border = border_thin

    # Deposit Info
    row += 2
    ws.cell(row, 1, f"訂金資訊: {deposit_info}")

    # Save file
    filename = f"Quotation_{quotation_no}.xlsx"
    filepath = os.path.join(output_path, filename)
    wb.save(filepath)
    return filename

def generate_pdf(company_name, company_address, company_phone, company_email, quotation_no, date, client_name, client_address, items, total_amount, received, balance, deposit_info):
    filename = f"Quotation_{quotation_no}.pdf"
    filepath = os.path.join(output_path, filename)
    doc = SimpleDocTemplate(filepath, pagesize=A4)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='SimSun', fontName='SimSun', fontSize=10, leading=14))
    styles.add(ParagraphStyle(name='SimSunBold', fontName='SimSun', fontSize=12, leading=14, alignment=1))
    styles.add(ParagraphStyle(name='SimSunTitle', fontName='SimSun', fontSize=16, leading=20, alignment=1))

    elements = []

    # Company Info
    elements.append(Paragraph(company_name, styles['SimSunTitle']))
    elements.append(Paragraph(company_address, styles['SimSunBold']))
    elements.append(Paragraph(f"Tel: {company_phone} Email: {company_email}", styles['SimSunBold']))
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph("報價單", styles['SimSunTitle']))
    elements.append(Spacer(1, 1*cm))

    # Client Info
    client_info_data = [
        [f"客戶名稱: {client_name}", f"報價單號碼: {quotation_no}"],
        [f"客戶地址: {client_address}", f"日期: {date}"]
    ]
    client_table = Table(client_info_data, colWidths=[10*cm, 6*cm])
    client_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'SimSun'),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 1, colors.black)
    ]))
    elements.append(client_table)
    elements.append(Spacer(1, 1*cm))

    # Items Table
    item_data = [["項目", "數量", "單價", "金額"]]
    for item in items:
        item_data.append([item['name'], item['quantity'], item['price'], item['amount']])
    
    item_table = Table(item_data, colWidths=[8*cm, 2*cm, 3*cm, 3*cm])
    item_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'SimSun'),
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 1, colors.black)
    ]))
    elements.append(item_table)

    # Total
    total_data = [
        ["", "", "總計:", total_amount],
        ["", "", "已收訂金:", received],
        ["", "", "餘額:", balance]
    ]
    total_table = Table(total_data, colWidths=[8*cm, 2*cm, 3*cm, 3*cm])
    total_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'SimSun'),
        ('ALIGN', (2,0), (2,2), 'RIGHT'),
        ('ALIGN', (3,0), (3,2), 'CENTER'),
        ('GRID', (2,0), (-1,-1), 1, colors.black)
    ]))
    elements.append(total_table)
    elements.append(Spacer(1, 1*cm))

    # Deposit Info
    elements.append(Paragraph(f"訂金資訊: {deposit_info}", styles['SimSun']))

    doc.build(elements)
    return filename

if __name__ == '__main__':
    app.run(debug=True)