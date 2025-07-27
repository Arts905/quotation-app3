import json
from flask import Flask, render_template, request, send_from_directory, jsonify
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import cm
import os

app = Flask(__name__)

# On serverless platforms like Vercel or Cloudflare Pages, use /tmp for writable files
if os.environ.get('VERCEL') or os.environ.get('CF_PAGES'):
    writable_dir = '/tmp'
else:
    # For local development, use the project directory
    writable_dir = os.path.abspath(os.path.dirname(__file__))

# Define paths for instance (database) and output (generated files)
instance_path = os.path.join(writable_dir, 'instance')
output_path = os.path.join(writable_dir, 'output')

# Create these directories if they don't exist
os.makedirs(instance_path, exist_ok=True)
os.makedirs(output_path, exist_ok=True)

# The base directory for reading non-writable files like fonts
# In Cloudflare Functions, the project root is the execution directory.
basedir = os.path.abspath(os.path.dirname(__file__))
# We need to go up one level from `functions` directory
basedir = os.path.dirname(basedir)


# Use DATABASE_URL from environment variables if set, otherwise use the writable path
default_db_path = os.path.join(instance_path, 'quotations.db')
database_url = os.environ.get('DATABASE_URL', f'sqlite:///{default_db_path}')

# Vercel's Postgres uses postgres:// which SQLAlchemy needs as postgresql://
if database_url and database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)

app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

class Quotation(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    company_name = db.Column(db.String(100))
    company_address = db.Column(db.String(200))
    company_phone = db.Column(db.String(20))
    company_email = db.Column(db.String(100))
    client_name = db.Column(db.String(100))
    client_address = db.Column(db.String(200))
    quotation_no = db.Column(db.String(50))
    date = db.Column(db.String(50))
    items = db.Column(db.Text)  # Store items as a JSON string
    received = db.Column(db.Float)
    deposit_info = db.Column(db.Text)

    def to_dict(self):
        return {
            'id': self.id,
            'company_name': self.company_name,
            'company_address': self.company_address,
            'company_phone': self.company_phone,
            'company_email': self.company_email,
            'client_name': self.client_name,
            'client_address': self.client_address,
            'quotation_no': self.quotation_no,
            'date': self.date,
            'items': self.items, # Keep as JSON string for loading
            'received': self.received,
            'deposit_info': self.deposit_info
        }

# Register Chinese font
# The font is expected to be in the 'static' directory relative to the project root
pdfmetrics.registerFont(TTFont('SimSun', os.path.join(basedir, 'static', 'SimSun.ttf')))

# with app.app_context():
#     db.create_all()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/quotations')
def get_quotations():
    quotations = Quotation.query.order_by(Quotation.date.desc()).all()
    return jsonify([q.to_dict() for q in quotations])

@app.route('/api/save_quotation', methods=['POST'])
def save_quotation():
    data = request.get_json()
    quotation_id = data.get('id')

    items_json = json.dumps(data.get('items', []))

    if quotation_id:
        # Update existing quotation
        quotation = Quotation.query.get(quotation_id)
        if not quotation:
            return jsonify({'success': False, 'message': 'Quotation not found'}), 404
    else:
        # Create new quotation
        quotation = Quotation()

    quotation.company_name = data.get('company_name')
    quotation.company_address = data.get('company_address')
    quotation.company_phone = data.get('company_phone')
    quotation.company_email = data.get('company_email')
    quotation.client_name = data.get('client_name')
    quotation.client_address = data.get('client_address')
    quotation.quotation_no = data.get('quotation_no')
    quotation.date = data.get('date')
    quotation.received = data.get('received')
    quotation.deposit_info = data.get('deposit_info')
    quotation.items = items_json

    if not quotation_id:
        db.session.add(quotation)
        
    db.session.commit()
    return jsonify({'success': True, 'id': quotation.id, 'message': 'Quotation saved successfully'})

@app.route('/api/get_quotation/<int:id>')
def get_quotation(id):
    quotation = Quotation.query.get(id)
    if quotation:
        return jsonify(quotation.to_dict())
    return jsonify({'error': 'Quotation not found'}), 404

@app.route('/api/delete_quotation/<int:id>', methods=['DELETE'])
def delete_quotation(id):
    quotation = Quotation.query.get(id)
    if quotation:
        db.session.delete(quotation)
        db.session.commit()
        return jsonify({'success': True})
    return jsonify({'error': 'Quotation not found'}), 404


@app.route('/generate', methods=['POST'])
def generate():
    # Collect data from form
    data = {
        'company_name': request.form['company_name'],
        'company_address': request.form['company_address'],
        'company_phone': request.form['company_phone'],
        'company_email': request.form['company_email'],
        'quotation_no': request.form['quotation_no'],
        'date': request.form['date'],
        'client_name': request.form['client_name'],
        'client_address': request.form['client_address'],
        'received': float(request.form.get('received', '0') or '0'),
        'deposit_info': request.form.get('deposit_info', ''),
        'items': []
    }

    item_names = request.form.getlist('item_name[]')
    quantities = request.form.getlist('quantity[]')
    prices = request.form.getlist('price[]')

    total_amount = 0
    for i in range(len(item_names)):
        if not item_names[i]: continue
        quantity = float(quantities[i])
        price = float(prices[i])
        amount = quantity * price
        data['items'].append({'name': item_names[i], 'quantity': quantity, 'price': price, 'amount': amount})
        total_amount += amount
    
    data['total_amount'] = total_amount
    data['balance'] = total_amount - data['received']
    items = data.pop('items')

    return render_template('preview.html', data=data, items=items)

@app.route('/create_files', methods=['POST'])
def create_files():
    # Get data from hidden fields
    company_name = request.form['company_name']
    company_address = request.form['company_address']
    company_phone = request.form['company_phone']
    company_email = request.form['company_email']
    quotation_no = request.form['quotation_no']
    date = request.form['date']
    client_name = request.form['client_name']
    client_address = request.form['client_address']
    received = float(request.form.get('received', '0') or '0')
    deposit_info = request.form.get('deposit_info', '')

    item_names = request.form.getlist('item_name[]')
    quantities = request.form.getlist('quantity[]')
    prices = request.form.getlist('price[]')

    items_data = []
    total_amount = 0
    for i in range(len(item_names)):
        if not item_names[i]: continue
        quantity = float(quantities[i])
        price = float(prices[i])
        amount = quantity * price
        items_data.append([item_names[i], quantity, f'{price:,.2f}', f'{amount:,.2f}'])
        total_amount += amount

    # The output directory is already created at the start of the app.
    # We just need to use the `output_path` variable.

    # Generate Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"

    # Styles
    font_bold = Font(bold=True)
    align_right = Alignment(horizontal='right')

    # Header
    ws.merge_cells('A1:B1')
    ws['A1'] = 'Quotation'
    ws['A1'].font = Font(size=20, bold=True)

    # Company Info
    ws['A3'] = company_name
    ws['A3'].font = font_bold
    ws['A4'] = company_address
    ws['A5'] = company_phone
    ws['A6'] = company_email

    # Quotation Info
    info_fill = PatternFill(start_color="f5f0e8", end_color="f5f0e8", fill_type="solid")
    for col in ['C', 'D']:
        for row in range(3, 7):
            ws[f'{col}{row}'].fill = info_fill

    ws['C3'] = 'Quotation No:'
    ws['D3'] = quotation_no
    ws['C4'] = 'Date:'
    ws['D4'] = date
    ws['C5'] = 'To:'
    ws['D5'] = client_name
    ws['C6'] = 'Address:'
    ws['D6'] = client_address

    # Items header
    headers = ['Items', 'Units', 'Cost Per Unit', 'Amount']
    header_fill = PatternFill(start_color="f5f0e8", end_color="f5f0e8", fill_type="solid")
    ws.append([]) # Spacer
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=col_num)
        cell.font = font_bold
        cell.fill = header_fill

    for item in items_data:
        ws.append(item)
    
    balance = total_amount - received
    summary_start_row = ws.max_row + 1
    ws.cell(row=summary_start_row, column=3, value='TOTAL').font = font_bold
    ws.cell(row=summary_start_row, column=4, value=f'{total_amount:,.2f}').alignment = align_right
    ws.cell(row=summary_start_row + 1, column=3, value='Received').font = font_bold
    ws.cell(row=summary_start_row + 1, column=4, value=f'({received:,.2f})').alignment = align_right
    ws.cell(row=summary_start_row + 2, column=3, value='Balance').font = font_bold
    ws.cell(row=summary_start_row + 2, column=4, value=f'{balance:,.2f}').alignment = align_right

    ws.cell(row=summary_start_row + 4, column=1, value=deposit_info)

    # Column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    excel_path = os.path.join(output_path, 'Quotation.xlsx')
    wb.save(excel_path)

    # Generate PDF
    pdf_path = os.path.join(output_path, 'Quotation.pdf')
    doc = SimpleDocTemplate(pdf_path, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm, leftMargin=2*cm, rightMargin=2*cm)
    
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='SimSun', fontName='SimSun', fontSize=10, leading=14))
    styles.add(ParagraphStyle(name='SimSunBold', fontName='SimSun', fontSize=10, leading=14, fontStyle='Bold'))
    styles.add(ParagraphStyle(name='SimSunRight', fontName='SimSun', fontSize=10, alignment=2))
    styles.add(ParagraphStyle(name='SimSunTitle', fontName='SimSun', fontSize=18, leading=22))

    story = []

    # Header
    header_data = [
        [Paragraph('<b>Quotation</b>', styles['SimSunTitle']), '', Paragraph(f'Quotation No: {quotation_no}', styles['SimSun']), ''],
        [Paragraph(company_name, styles['SimSunBold']), '', Paragraph(f'Date: {date}', styles['SimSun']), ''],
        [Paragraph(company_address, styles['SimSun']), '', Paragraph(f'To: {client_name}', styles['SimSun']), ''],
        [Paragraph(company_phone, styles['SimSun']), '', Paragraph(f'Address: {client_address}', styles['SimSun']), ''],
        [Paragraph(company_email, styles['SimSun']), '', '', '']
    ]
    header_table = Table(header_data, colWidths=[7*cm, 3*cm, 7*cm, 0*cm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('SPAN', (0,0), (1,0)),
        ('SPAN', (2,0), (3,0)),
        ('SPAN', (2,1), (3,1)),
        ('SPAN', (2,2), (3,2)),
        ('SPAN', (2,3), (3,3)),
        ('BACKGROUND', (2,0), (3,4), colors.HexColor('#f5f0e8')),
        ('FONTNAME', (2,0), (3,4), 'SimSun'),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 1*cm))

    # Items Table
    items_header = [Paragraph(h, styles['SimSunBold']) for h in ['Items', 'Units', 'Cost Per Unit', 'Amount']]
    items_list = [items_header] + [[Paragraph(str(col), styles['SimSun']) for col in row] for row in items_data]
    
    items_table = Table(items_list, colWidths=[8*cm, 2*cm, 3*cm, 4*cm])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f5f0e8')), # Header color
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#fafafa')), # Item rows color
        ('TEXTCOLOR',(0,0),(-1,0),colors.black),
        ('ALIGN', (1,1), (-1,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#dddddd')) # A more visible grid
    ]))
    story.append(items_table)

    # Summary
    balance = total_amount - received
    summary_data = [
        ['', 'TOTAL', f'{total_amount:,.2f}'],
        ['', 'Received', f'({received:,.2f})'],
        ['', 'Balance', f'{balance:,.2f}']
    ]
    summary_table = Table(summary_data, colWidths=[11*cm, 2*cm, 4*cm])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
        ('FONTNAME', (1,0), (1,-1), 'SimSun'),
        ('FONTNAME', (2,0), (2,-1), 'SimSun'),
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f5f0e8')),
        ('GRID', (1,0), (-1,-1), 1, colors.white)
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 1*cm))

    # Footer
    story.append(Paragraph(deposit_info, styles['SimSun']))
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph('Thank you for your business!', styles['SimSunRight']))

    doc.build(story)

    return render_template('result.html', excel_path='Quotation.xlsx', pdf_path='Quotation.pdf')

@app.route('/download/<filename>')
def download(filename):
    return send_from_directory(output_path, filename, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)